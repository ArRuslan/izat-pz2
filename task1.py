from pathlib import Path

import click
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill

PATH_PARAM = click.Path(readable=True, resolve_path=True, dir_okay=False)
PATH_PARAM_EXISTS = click.Path(exists=True, readable=True, resolve_path=True, dir_okay=False)

green = "8000FF00"
blue = "800000FF"
yellow = "80FFFF00"

fill_green = PatternFill(start_color=green, end_color=green, fill_type="solid")
fill_blue = PatternFill(start_color=blue, end_color=blue, fill_type="solid")
fill_yellow = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")


def _process_cell(cell: Cell) -> None:
    if not isinstance(cell.value, (int, str)):
        return

    if isinstance(cell.value, int):
        cell.fill = fill_green
        return

    if cell.value.isdigit():
        cell.fill = fill_green
        return

    chars = False
    digits = False
    for char in cell.value:
        if char.isdigit():
            digits = True
        else:
            chars = True

    if chars and digits:
        cell.fill = fill_yellow
    elif chars:
        cell.fill = fill_blue
    elif digits:
        cell.fill = fill_green


@click.command()
@click.option("--input-file", "-i", type=PATH_PARAM_EXISTS, required=True)
@click.option("--output-file", "-o", type=PATH_PARAM, required=True)
def main(input_file: Path, output_file: Path) -> None:
    wb = load_workbook(input_file)
    ws = wb.active

    for row in ws:
        for col in row:
            _process_cell(col)

    wb.save(output_file)


if __name__ == "__main__":
    main()
